import os
import sys
import io
from datetime import datetime, timezone, timedelta

from fastapi import FastAPI, HTTPException, Depends, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from sqlalchemy.orm import Session, joinedload

from database import engine, get_db, Base
from models import Drink, DrinkCategory, Order, OrderItem
from schemas import (
    CategoryOut, DrinkOut, DrinkCreate, DrinkUpdate,
    OrderCreate, OrderOut, OrderItemOut, StatusUpdate, StatsOut,
    DashboardOut, DailyPoint, TopDrink, CategorySlice, HourlyPoint,
)
from seed import seed

app = FastAPI(title="Bar Order System")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    seed()


# ── Category ──
@app.get("/api/categories", response_model=list[CategoryOut])
def list_categories(db: Session = Depends(get_db)):
    return db.query(DrinkCategory).order_by(DrinkCategory.sort_order).all()


# ── Drinks ──
@app.get("/api/drinks", response_model=list[DrinkOut])
def list_drinks(category_id: int | None = None, search: str = "",
                db: Session = Depends(get_db)):
    q = db.query(Drink)
    if category_id:
        q = q.filter(Drink.category_id == category_id)
    if search:
        q = q.filter(Drink.name.contains(search))
    return q.order_by(Drink.name).all()


@app.post("/api/drinks", response_model=DrinkOut)
def create_drink(data: DrinkCreate, db: Session = Depends(get_db)):
    cat = db.query(DrinkCategory).filter(DrinkCategory.id == data.category_id).first()
    if not cat:
        raise HTTPException(404, "分类不存在")
    drink = Drink(**data.model_dump())
    db.add(drink)
    db.commit()
    db.refresh(drink)
    return drink


@app.put("/api/drinks/{drink_id}", response_model=DrinkOut)
def update_drink(drink_id: int, data: DrinkUpdate, db: Session = Depends(get_db)):
    drink = db.query(Drink).filter(Drink.id == drink_id).first()
    if not drink:
        raise HTTPException(404, "酒款不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(drink, k, v)
    db.commit()
    db.refresh(drink)
    return drink


@app.delete("/api/drinks/{drink_id}")
def delete_drink(drink_id: int, db: Session = Depends(get_db)):
    drink = db.query(Drink).filter(Drink.id == drink_id).first()
    if not drink:
        raise HTTPException(404, "酒款不存在")
    db.delete(drink)
    db.commit()
    return {"ok": True}


# ── Orders ──
@app.post("/api/orders", response_model=OrderOut)
def create_order(data: OrderCreate, db: Session = Depends(get_db)):
    if not data.items:
        raise HTTPException(400, "订单至少需要一个商品")

    order = Order(table_number=data.table_number, note=data.note, total_price=0)
    db.add(order)
    db.flush()

    total = 0
    items = []
    for it in data.items:
        drink = db.query(Drink).filter(Drink.id == it.drink_id).first()
        if not drink:
            raise HTTPException(400, f"酒款 {it.drink_id} 不存在")
        if not drink.is_available:
            raise HTTPException(400, f"「{drink.name}」已下架")
        if drink.stock < it.quantity:
            raise HTTPException(400, f"「{drink.name}」库存不足 (剩余 {drink.stock})")

        drink.stock -= it.quantity
        item = OrderItem(order_id=order.id, drink_id=drink.id,
                         quantity=it.quantity, unit_price=drink.price)
        db.add(item)
        total += drink.price * it.quantity
        items.append(item)

    order.total_price = total
    db.commit()
    db.refresh(order)

    result = _order_to_out(order)
    return result


@app.get("/api/orders", response_model=list[OrderOut])
def list_orders(status: str | None = None, table: str = "",
                db: Session = Depends(get_db)):
    q = db.query(Order).options(joinedload(Order.items))
    if status:
        q = q.filter(Order.status == status)
    if table:
        q = q.filter(Order.table_number.contains(table))
    q = q.order_by(Order.created_at.desc())
    return [_order_to_out(o) for o in q.all()]


@app.get("/api/orders/export")
def export_orders(
    date: str | None = Query(default=None, description="日期 YYYY-MM-DD"),
    status: str | None = Query(default=None, description="订单状态"),
    db: Session = Depends(get_db),
):
    q = db.query(Order).options(joinedload(Order.items))
    if date:
        try:
            d = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(400, "日期格式错误，应为 YYYY-MM-DD")
        q = q.filter(Order.created_at >= d, Order.created_at < d + timedelta(days=1))
    if status:
        q = q.filter(Order.status == status)
    orders = q.order_by(Order.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"

    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(vertical="center")
    money_fmt = "#,##0.00"
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    total_fill = PatternFill(start_color="FFF3D4", end_color="FFF3D4", fill_type="solid")

    headers = ["订单号", "桌号", "状态", "下单时间", "酒款名称", "数量", "单价", "金额", "备注"]
    col_widths = [10, 10, 10, 22, 32, 8, 10, 12, 24]

    title = f"酒吧点单明细 - {date or '全部日期'}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title).font = Font(name="Microsoft YaHei", bold=True, size=14, color="D4A853")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 22

    row = 3
    total_all = 0
    status_labels = {"pending": "待制作", "in_progress": "制作中", "done": "已完成", "cancelled": "已取消"}
    for order in orders:
        total_all += order.total_price
        for i, item in enumerate(order.items):
            drink_name = item.drink.name if item.drink else "(已删除)"
            ws.cell(row=row, column=1, value=order.id).border = thin_border
            ws.cell(row=row, column=2, value=order.table_number).border = thin_border
            ws.cell(row=row, column=3, value=status_labels.get(order.status, order.status)).border = thin_border
            ws.cell(row=row, column=4, value=order.created_at.strftime("%Y-%m-%d %H:%M")).border = thin_border
            ws.cell(row=row, column=5, value=drink_name).border = thin_border
            ws.cell(row=row, column=6, value=item.quantity).border = thin_border
            c_price = ws.cell(row=row, column=7, value=item.unit_price)
            c_price.number_format = money_fmt
            c_price.border = thin_border
            c_amt = ws.cell(row=row, column=8, value=item.quantity * item.unit_price)
            c_amt.number_format = money_fmt
            c_amt.border = thin_border
            ws.cell(row=row, column=9, value=order.note if i == 0 else "").border = thin_border
            for c in range(1, 10):
                ws.cell(row=row, column=c).alignment = cell_align
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        subtotal_cell = ws.cell(row=row, column=1, value=f"订单 #{order.id} 合计")
        subtotal_cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
        subtotal_cell.alignment = Alignment(horizontal="right", vertical="center")
        subtotal_cell.fill = total_fill
        for c in range(1, 10):
            ws.cell(row=row, column=c).fill = total_fill
            ws.cell(row=row, column=c).border = thin_border
        amt_cell = ws.cell(row=row, column=8, value=order.total_price)
        amt_cell.font = Font(name="Microsoft YaHei", bold=True, size=10, color="D4A853")
        amt_cell.number_format = money_fmt
        amt_cell.fill = total_fill
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    gt_cell = ws.cell(row=row, column=1, value="总计")
    gt_cell.font = Font(name="Microsoft YaHei", bold=True, size=12)
    gt_cell.alignment = Alignment(horizontal="right", vertical="center")
    for c in range(1, 10):
        ws.cell(row=row, column=c).font = Font(name="Microsoft YaHei", bold=True, size=12)
        ws.cell(row=row, column=c).border = Border(
            top=Side(style="medium", color="D4A853"),
            bottom=Side(style="double", color="D4A853"),
        )
    gt_amt_cell = ws.cell(row=row, column=8, value=total_all)
    gt_amt_cell.number_format = money_fmt
    gt_amt_cell.font = Font(name="Microsoft YaHei", bold=True, size=12, color="D4A853")

    ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"orders_{date or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/orders/{order_id}", response_model=OrderOut)
def get_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).options(joinedload(Order.items)).filter(
        Order.id == order_id).first()
    if not order:
        raise HTTPException(404, "订单不存在")
    return _order_to_out(order)


@app.patch("/api/orders/{order_id}/status", response_model=OrderOut)
def update_order_status(order_id: int, data: StatusUpdate,
                        db: Session = Depends(get_db)):
    allowed = {"pending", "in_progress", "done", "cancelled"}
    if data.status not in allowed:
        raise HTTPException(400, f"无效状态，可选: {allowed}")

    order = db.query(Order).options(joinedload(Order.items)).filter(
        Order.id == order_id).first()
    if not order:
        raise HTTPException(404, "订单不存在")

    if data.status == "cancelled" and order.status != "cancelled":
        for item in order.items:
            drink = db.query(Drink).filter(Drink.id == item.drink_id).first()
            if drink:
                drink.stock += item.quantity

    order.status = data.status
    order.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(order)
    return _order_to_out(order)


# ── Stats ──
@app.get("/api/stats", response_model=StatsOut)
def get_stats(db: Session = Depends(get_db)):
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_orders = db.query(Order).filter(
        Order.created_at >= today, Order.status != "cancelled").all()
    return StatsOut(
        today_orders=len(today_orders),
        today_revenue=sum(o.total_price for o in today_orders),
        pending_count=db.query(Order).filter(Order.status == "pending").count(),
        in_progress_count=db.query(Order).filter(Order.status == "in_progress").count(),
    )


# ── Analytics ──
@app.get("/api/analytics/dashboard", response_model=DashboardOut)
def analytics_dashboard(days: int = 30, db: Session = Depends(get_db)):
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    orders = db.query(Order).options(joinedload(Order.items)).filter(
        Order.created_at >= cutoff,
        Order.status != "cancelled"
    ).all()

    # Daily trend
    daily_map: dict[str, dict] = {}
    for o in orders:
        dk = o.created_at.strftime("%Y-%m-%d")
        if dk not in daily_map:
            daily_map[dk] = {"revenue": 0, "orders": 0}
        daily_map[dk]["revenue"] += o.total_price
        daily_map[dk]["orders"] += 1
    daily_trend = []
    for i in range(days):
        d = (datetime.now(timezone.utc) - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
        pt = daily_map.get(d, {"revenue": 0, "orders": 0})
        daily_trend.append(DailyPoint(date=d, revenue=round(pt["revenue"], 2), orders=pt["orders"]))

    # Top drinks
    drink_sales: dict[int, dict] = {}
    for o in orders:
        for item in o.items:
            did = item.drink_id
            if did not in drink_sales:
                drink_sales[did] = {"qty": 0, "rev": 0.0}
            drink_sales[did]["qty"] += item.quantity
            drink_sales[did]["rev"] += item.quantity * item.unit_price
    top = sorted(drink_sales.items(), key=lambda x: x[1]["qty"], reverse=True)[:10]
    drink_ids = [t[0] for t in top]
    drinks_map = {d.id: d for d in db.query(Drink).filter(Drink.id.in_(drink_ids)).all()}
    top_drinks = []
    for did, data in top:
        d = drinks_map.get(did)
        top_drinks.append(TopDrink(
            drink_id=did,
            name=d.name if d else "(已删除)",
            category_name=d.category.name if d and d.category else "",
            quantity=data["qty"],
            revenue=round(data["rev"], 2),
        ))

    # Category breakdown
    cat_revenue: dict[str, float] = {}
    total_rev = 0.0
    for o in orders:
        for item in o.items:
            d = drinks_map.get(item.drink_id) or db.query(Drink).filter(Drink.id == item.drink_id).first()
            cat = d.category.name if d and d.category else "未分类"
            rev = item.quantity * item.unit_price
            cat_revenue[cat] = cat_revenue.get(cat, 0) + rev
            total_rev += rev
    category_breakdown = [
        CategorySlice(category_name=cat, revenue=round(rev, 2),
                      percentage=round(rev / total_rev * 100, 1) if total_rev > 0 else 0)
        for cat, rev in sorted(cat_revenue.items(), key=lambda x: x[1], reverse=True)
    ]

    # Hourly distribution
    hourly: dict[int, int] = {h: 0 for h in range(24)}
    for o in orders:
        hourly[o.created_at.hour] += 1
    hourly_distribution = [HourlyPoint(hour=h, orders=c) for h, c in hourly.items()]

    total_revenue = sum(o.total_price for o in orders)
    total_orders = len(orders)

    return DashboardOut(
        total_revenue=round(total_revenue, 2),
        total_orders=total_orders,
        avg_order_value=round(total_revenue / total_orders, 2) if total_orders > 0 else 0,
        daily_trend=daily_trend,
        top_drinks=top_drinks,
        category_breakdown=category_breakdown,
        hourly_distribution=hourly_distribution,
    )


class ResetBody(BaseModel):
    mode: str = "orders"


@app.post("/api/reset")
def reset_data(body: ResetBody, db: Session = Depends(get_db)):
    mode = body.mode
    if mode == "orders":
        # Restore stock for all non-cancelled order items
        orders = db.query(Order).filter(Order.status != "cancelled").all()
        for order in orders:
            for item in order.items:
                drink = db.query(Drink).filter(Drink.id == item.drink_id).first()
                if drink:
                    drink.stock += item.quantity
        # Delete all order items and orders
        db.query(OrderItem).delete()
        db.query(Order).delete()
        db.commit()
        return {"ok": True, "message": "已清除所有订单，库存已恢复"}

    if mode == "full":
        # Drop and recreate all tables, then seed
        Base.metadata.drop_all(bind=engine)
        Base.metadata.create_all(bind=engine)
        seed()
        return {"ok": True, "message": "已完全重置，数据恢复为初始状态"}

    raise HTTPException(400, "mode 必须为 orders 或 full")


@app.get("/api/health")
def health():
    return {"status": "ok"}


if getattr(sys, "frozen", False):
    # PyInstaller bundle
    _BASE = sys._MEIPASS
else:
    _BASE = os.path.join(os.path.dirname(__file__), "..")

FRONTEND_DIST = os.path.join(_BASE, "frontend", "dist")


@app.get("/{full_path:path}")
async def serve_frontend(full_path: str):
    if not os.path.isdir(FRONTEND_DIST):
        raise HTTPException(404, "前端未构建")
    file_path = os.path.join(FRONTEND_DIST, full_path)
    if os.path.isfile(file_path):
        return FileResponse(file_path)
    # SPA fallback: all non-file routes serve index.html
    index = os.path.join(FRONTEND_DIST, "index.html")
    if os.path.isfile(index):
        return FileResponse(index)
    raise HTTPException(404)


def _order_to_out(order: Order) -> OrderOut:
    items = []
    for item in order.items:
        items.append(OrderItemOut(
            id=item.id,
            drink_id=item.drink_id,
            drink_name=item.drink.name if item.drink else "",
            quantity=item.quantity,
            unit_price=item.unit_price,
        ))
    return OrderOut(
        id=order.id,
        table_number=order.table_number,
        status=order.status,
        total_price=order.total_price,
        note=order.note,
        created_at=order.created_at,
        updated_at=order.updated_at,
        items=items,
    )
